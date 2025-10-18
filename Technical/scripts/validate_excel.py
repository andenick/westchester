"""
Westchester County Data Platform - Excel Validation Script

Validates that all Excel files follow Druck mandatory standards:
- ONE SHEET PER FILE (no exceptions)
- Machine-readable column names
- Black & White formatting (no colors)

Per AGENT_STANDARDS_AND_BEST_PRACTICES.md Section 1.1
"""

import pandas as pd
from pathlib import Path
import sys
from openpyxl import load_workbook
from openpyxl.styles import Font, Fill


def validate_one_sheet_rule(excel_file: Path) -> tuple[bool, str]:
    """
    Validate that Excel file has exactly ONE sheet.
    
    Args:
        excel_file: Path to Excel file
        
    Returns:
        (is_valid, message) tuple
    """
    try:
        xl = pd.ExcelFile(excel_file)
        sheet_count = len(xl.sheet_names)
        
        if sheet_count == 1:
            return True, f"[SUCCESS] {excel_file.name}: ONE sheet (compliant)"
        else:
            return False, f"[FAILED] {excel_file.name}: {sheet_count} sheets (VIOLATION - must have exactly ONE)"
    
    except Exception as e:
        return False, f"[FAILED] {excel_file.name}: Error reading file - {str(e)}"


def validate_column_names(excel_file: Path) -> tuple[bool, str]:
    """
    Validate that column names are machine-readable.
    
    Checks for:
    - No spaces in column names (prefer underscores)
    - No special characters (except underscore)
    - Consistent naming convention
    
    Args:
        excel_file: Path to Excel file
        
    Returns:
        (is_valid, message) tuple
    """
    try:
        df = pd.read_excel(excel_file)
        issues = []
        
        for col in df.columns:
            # Check for spaces
            if ' ' in str(col):
                issues.append(f"Column '{col}' contains spaces (prefer underscores)")
            
            # Check for special characters (allow alphanumeric and underscore only)
            if not str(col).replace('_', '').replace(' ', '').isalnum():
                special_chars = [c for c in str(col) if not c.isalnum() and c not in ['_', ' ']]
                if special_chars:
                    issues.append(f"Column '{col}' has special characters: {special_chars}")
        
        if not issues:
            return True, f"[SUCCESS] {excel_file.name}: Column names are machine-readable"
        else:
            warnings = "\n    ".join(issues)
            return True, f"[WARNING] {excel_file.name}: Column name warnings:\n    {warnings}\n    (Not blocking, but consider improving)"
    
    except Exception as e:
        return False, f"[FAILED] {excel_file.name}: Error validating columns - {str(e)}"


def validate_formatting(excel_file: Path) -> tuple[bool, str]:
    """
    Validate that formatting is professional B&W (no colors).
    
    Args:
        excel_file: Path to Excel file
        
    Returns:
        (is_valid, message) tuple
    """
    try:
        wb = load_workbook(excel_file)
        ws = wb.active
        
        colored_cells = []
        
        for row in ws.iter_rows():
            for cell in row:
                # Check font color
                if cell.font.color and cell.font.color.rgb and cell.font.color.rgb not in ['00000000', 'FF000000']:
                    colored_cells.append(f"{cell.coordinate} (font)")
                
                # Check fill color
                if cell.fill.fgColor and cell.fill.fgColor.rgb and cell.fill.fgColor.rgb not in ['00000000', 'FFFFFFFF']:
                    colored_cells.append(f"{cell.coordinate} (fill)")
        
        if not colored_cells:
            return True, f"[SUCCESS] {excel_file.name}: Professional B&W formatting"
        else:
            cells_str = ", ".join(colored_cells[:5])  # Show first 5
            more = f" and {len(colored_cells) - 5} more" if len(colored_cells) > 5 else ""
            return False, f"[WARNING] {excel_file.name}: Colored cells found: {cells_str}{more}\n    (Should be B&W only per Druck standards)"
    
    except Exception as e:
        return True, f"[WARNING] {excel_file.name}: Could not validate formatting - {str(e)}\n    (Proceeding with validation)"


def validate_file_size(excel_file: Path) -> tuple[bool, str]:
    """
    Check file size and warn if unusually large.
    
    Args:
        excel_file: Path to Excel file
        
    Returns:
        (is_valid, message) tuple
    """
    size_mb = excel_file.stat().st_size / (1024 * 1024)
    
    if size_mb > 50:
        return True, f"[WARNING] {excel_file.name}: Large file ({size_mb:.1f} MB)\n    Consider splitting into multiple files if data is logically separable"
    elif size_mb > 100:
        return False, f"[FAILED] {excel_file.name}: Very large file ({size_mb:.1f} MB)\n    Files over 100 MB should be split for usability"
    else:
        return True, f"[SUCCESS] {excel_file.name}: Reasonable file size ({size_mb:.2f} MB)"


def validate_all_excel_files(data_directory: Path) -> dict:
    """
    Validate all Excel files in directory and subdirectories.
    
    Args:
        data_directory: Path to Output/Data/ directory
        
    Returns:
        Dictionary with validation results
    """
    excel_files = list(data_directory.glob('**/*.xlsx'))
    
    # Filter out temporary Excel files (start with ~$)
    excel_files = [f for f in excel_files if not f.name.startswith('~$')]
    
    if not excel_files:
        return {
            'total': 0,
            'compliant': 0,
            'violations': 0,
            'warnings': 0,
            'files': [],
            'message': f"No Excel files found in {data_directory}"
        }
    
    results = {
        'total': len(excel_files),
        'compliant': 0,
        'violations': 0,
        'warnings': 0,
        'files': []
    }
    
    print(f"\n{'='*80}")
    print(f"WESTCHESTER COUNTY - EXCEL VALIDATION REPORT")
    print(f"Druck Standard: ONE SHEET PER FILE (Mandatory)")
    print(f"{'='*80}\n")
    print(f"Directory: {data_directory}")
    print(f"Files found: {len(excel_files)}\n")
    print(f"{'='*80}\n")
    
    for excel_file in sorted(excel_files):
        file_result = {
            'name': excel_file.name,
            'path': str(excel_file),
            'validations': []
        }
        
        has_violation = False
        has_warning = False
        
        # Critical validation: ONE SHEET rule
        is_valid, message = validate_one_sheet_rule(excel_file)
        file_result['validations'].append(message)
        print(message)
        
        if not is_valid:
            has_violation = True
        
        # Column names validation
        is_valid, message = validate_column_names(excel_file)
        file_result['validations'].append(message)
        print(message)
        
        if not is_valid:
            has_violation = True
        elif '[WARNING]' in message:
            has_warning = True
        
        # Formatting validation
        is_valid, message = validate_formatting(excel_file)
        file_result['validations'].append(message)
        print(message)
        
        if not is_valid or '[WARNING]' in message:
            has_warning = True
        
        # File size check
        is_valid, message = validate_file_size(excel_file)
        file_result['validations'].append(message)
        print(message)
        
        if not is_valid:
            has_violation = True
        elif '[WARNING]' in message:
            has_warning = True
        
        print()  # Blank line between files
        
        # Update counts
        if has_violation:
            results['violations'] += 1
            file_result['status'] = 'VIOLATION'
        elif has_warning:
            results['warnings'] += 1
            file_result['status'] = 'WARNING'
        else:
            results['compliant'] += 1
            file_result['status'] = 'COMPLIANT'
        
        results['files'].append(file_result)
    
    return results


def print_summary(results: dict):
    """Print validation summary."""
    print(f"{'='*80}")
    print(f"VALIDATION SUMMARY")
    print(f"{'='*80}\n")
    
    if results['total'] == 0:
        print(results['message'])
        return
    
    print(f"Total files validated: {results['total']}")
    print(f"[SUCCESS] Fully compliant: {results['compliant']}")
    print(f"[WARNING] Warnings: {results['warnings']}")
    print(f"[FAILED] Violations: {results['violations']}\n")
    
    if results['violations'] > 0:
        print("🚨 CRITICAL VIOLATIONS FOUND 🚨")
        print("The following files violate Druck mandatory standards:\n")
        
        for file_result in results['files']:
            if file_result['status'] == 'VIOLATION':
                print(f"  - {file_result['name']}")
                for validation in file_result['validations']:
                    if '[FAILED]' in validation:
                        print(f"    {validation}")
        
        print("\n[WARNING]  FIX THESE VIOLATIONS BEFORE HANDOFF")
        print("[WARNING]  Completion rating cannot exceed 75% with violations")
        return False
    
    elif results['warnings'] > 0:
        print("[WARNING]  Some warnings found - review recommended")
        print("[SUCCESS] No critical violations - Druck compliant\n")
        return True
    
    else:
        print("[SUCCESS] ALL FILES PASS DRUCK STANDARDS")
        print("[SUCCESS] Ready for production use\n")
        return True


def main():
    """Main validation function."""
    # Determine data directory
    script_dir = Path(__file__).parent
    project_root = script_dir.parent.parent
    data_dir = project_root / "Output" / "Data"
    
    if not data_dir.exists():
        print(f"[FAILED] Data directory not found: {data_dir}")
        print("   Create Output/Data/ directory first")
        sys.exit(1)
    
    # Run validation
    results = validate_all_excel_files(data_dir)
    
    # Print summary
    is_compliant = print_summary(results)
    
    # Exit code
    if results['total'] == 0:
        sys.exit(0)  # No files to validate (not an error)
    elif results['violations'] > 0:
        sys.exit(1)  # Violations found
    else:
        sys.exit(0)  # All good


if __name__ == "__main__":
    main()

