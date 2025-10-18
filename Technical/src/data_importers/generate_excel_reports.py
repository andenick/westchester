"""
Generate Druck-Compliant Excel Reports for Sidewalk Coverage Analysis

CRITICAL: Each Excel file must have ONE SHEET ONLY (Druck standard)

Purpose: Create Excel outputs that answer Taylor's question front and center
with supporting data below.

Input: JSON statistics from county-wide analysis
Output: 4 Excel files, each with ONE SHEET
"""

import pandas as pd
import json
from pathlib import Path
from datetime import datetime

class ExcelReportGenerator:
    def __init__(self, base_dir):
        self.base_dir = Path(base_dir)
        self.analysis_dir = self.base_dir / "data" / "processed" / "countywide_sidewalk_analysis"
        self.output_dir = self.base_dir.parent / "Output" / "Excel"
        self.output_dir.mkdir(parents=True, exist_ok=True)

        # Load statistics
        with open(self.analysis_dir / "county_wide_statistics.json", 'r') as f:
            self.county_stats = json.load(f)

        with open(self.analysis_dir / "tod_statistics.json", 'r') as f:
            self.tod_stats = json.load(f)

    def generate_executive_summary(self):
        """
        ONE SHEET: Executive summary with answer front and center
        """
        print("\n=== Generating Executive Summary Excel (ONE SHEET) ===")

        # Extract key statistics
        tod_coverage = self.tod_stats['tod_statistics']['any_coverage_pct']
        tod_total = self.tod_stats['tod_statistics']['total_roads']
        tod_with_sidewalks = (self.tod_stats['tod_statistics']['one_side'] +
                             self.tod_stats['tod_statistics']['both_sides'])
        tod_without = self.tod_stats['tod_statistics']['no_coverage']

        county_coverage = self.county_stats['coverage_percentages']['any_coverage_pct']
        county_total = self.county_stats['total_roads']
        county_no_coverage = self.county_stats['coverage_counts']['no_coverage']

        non_tod_coverage = self.tod_stats['non_tod_statistics']['any_coverage_pct']

        # Create DataFrame with answer front and center
        data = []

        # HEADER ROW
        data.append(['SIDEWALK COVERAGE ASSESSMENT - WESTCHESTER COUNTY', '', '', ''])
        data.append(['Analysis Date:', datetime.now().strftime('%Y-%m-%d'), '', ''])
        data.append(['', '', '', ''])

        # ANSWER TO TAYLOR'S QUESTION - FRONT AND CENTER
        data.append(['═' * 80, '', '', ''])
        data.append(['ANSWER: METRO-NORTH STATION AREA (0.5 MILE) SIDEWALK COVERAGE', '', '', ''])
        data.append(['═' * 80, '', '', ''])
        data.append(['', '', '', ''])
        data.append(['TOD Area Coverage:', f'{tod_coverage}%', '', 'MODERATE ADEQUACY'])
        data.append(['Roads with Sidewalks:', tod_with_sidewalks, f'out of {tod_total} total', ''])
        data.append(['Roads WITHOUT Sidewalks:', tod_without, f'({100-tod_coverage:.1f}% lack coverage)', ''])
        data.append(['', '', '', ''])

        # COMPARISON TO COUNTY AVERAGE
        data.append(['═' * 80, '', '', ''])
        data.append(['COMPARISON: TOD vs COUNTY-WIDE', '', '', ''])
        data.append(['═' * 80, '', '', ''])
        data.append(['', '', '', ''])
        data.append(['Area Type', 'Coverage %', 'Roads with Sidewalks', 'Total Roads'])
        data.append(['Metro-North Station Areas (TOD)', f'{tod_coverage}%', tod_with_sidewalks, tod_total])
        data.append(['County-Wide Average', f'{county_coverage}%',
                    county_total - county_no_coverage, county_total])
        data.append(['Non-TOD Areas', f'{non_tod_coverage}%', '', ''])
        data.append(['', '', '', ''])
        data.append(['Key Finding:', f'TOD areas have {tod_coverage/non_tod_coverage:.1f}x better coverage than non-TOD areas', '', ''])
        data.append(['', '', '', ''])

        # DETAILED BREAKDOWN
        data.append(['═' * 80, '', '', ''])
        data.append(['DETAILED BREAKDOWN - TOD AREA ROADS', '', '', ''])
        data.append(['═' * 80, '', '', ''])
        data.append(['', '', '', ''])
        data.append(['Coverage Type', 'Road Count', 'Percentage', 'Description'])
        data.append(['No Coverage',
                    self.tod_stats['tod_statistics']['no_coverage'],
                    f"{self.tod_stats['tod_statistics']['no_coverage']/tod_total*100:.1f}%",
                    'No sidewalks on either side'])
        data.append(['One-Side Coverage',
                    self.tod_stats['tod_statistics']['one_side'],
                    f"{self.tod_stats['tod_statistics']['one_side']/tod_total*100:.1f}%",
                    'Sidewalk on one side only'])
        data.append(['Both-Sides Coverage',
                    self.tod_stats['tod_statistics']['both_sides'],
                    f"{self.tod_stats['tod_statistics']['both_sides']/tod_total*100:.1f}%",
                    'Sidewalks on both sides'])
        data.append(['', '', '', ''])

        # METHODOLOGY NOTE
        data.append(['═' * 80, '', '', ''])
        data.append(['METHODOLOGY', '', '', ''])
        data.append(['═' * 80, '', '', ''])
        data.append(['Analysis Method:', 'DVRPC Sidewalk-to-Road Ratio (City Planning Standard)', '', ''])
        data.append(['Buffer Distance:', '0.5 miles (2,640 feet) from Metro-North stations', '', ''])
        data.append(['Coordinate System:', 'EPSG:2260 (NY State Plane Long Island, feet)', '', ''])
        data.append(['Data Source:', "Taylor's County Shapefiles", '', ''])
        data.append(['Roads Analyzed:', f'{county_total:,} county road polygons', '', ''])
        data.append(['Road Area Analyzed:', f'{self.county_stats["total_road_area_acres"]:,.1f} acres', '', ''])

        # Create DataFrame
        df = pd.DataFrame(data)

        # Write to Excel (ONE SHEET ONLY)
        output_file = self.output_dir / "1_EXECUTIVE_SUMMARY.xlsx"
        with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='Executive Summary', index=False, header=False)

            # Format the worksheet
            worksheet = writer.sheets['Executive Summary']

            # Set column widths
            worksheet.column_dimensions['A'].width = 50
            worksheet.column_dimensions['B'].width = 25
            worksheet.column_dimensions['C'].width = 30
            worksheet.column_dimensions['D'].width = 25

            # Bold the answer rows
            from openpyxl.styles import Font, Alignment, PatternFill

            for row in range(1, len(data) + 1):
                cell = worksheet.cell(row, 1)
                if row in [1, 5, 6, 8]:  # Header and answer rows
                    cell.font = Font(bold=True, size=12)
                if row == 8:  # Main answer
                    cell.font = Font(bold=True, size=14)
                    cell.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

        print(f"[OK] Created: {output_file}")
        print(f"  Answer front and center: TOD Coverage = {tod_coverage}%")
        return output_file

    def generate_tod_comparison(self):
        """
        ONE SHEET: TOD vs Non-TOD detailed comparison
        """
        print("\n=== Generating TOD Comparison Excel (ONE SHEET) ===")

        data = []

        # Header
        data.append(['TRANSIT-ORIENTED DEVELOPMENT (TOD) COVERAGE COMPARISON', '', '', ''])
        data.append(['Metro-North Station Areas vs Rest of County', '', '', ''])
        data.append(['', '', '', ''])

        # Side-by-side comparison
        data.append(['Metric', 'TOD Areas (0.5 mi)', 'Non-TOD Areas', 'Difference'])
        data.append(['', '', '', ''])

        tod = self.tod_stats['tod_statistics']
        non_tod = self.tod_stats['non_tod_statistics']

        data.append(['ANY Coverage Percentage',
                    f"{tod['any_coverage_pct']:.1f}%",
                    f"{non_tod['any_coverage_pct']:.1f}%",
                    f"+{tod['any_coverage_pct'] - non_tod['any_coverage_pct']:.1f}%"])
        data.append(['Total Roads', tod['total_roads'], non_tod['total_roads'], ''])
        data.append(['', '', '', ''])

        # Coverage breakdown
        data.append(['Coverage Type Breakdown', 'TOD Areas', 'Non-TOD Areas', ''])
        data.append(['No Coverage', tod['no_coverage'], non_tod['no_coverage'], ''])
        data.append(['One-Side Coverage', tod['one_side'], non_tod['one_side'], ''])
        data.append(['Both-Sides Coverage', tod['both_sides'], non_tod['both_sides'], ''])
        data.append(['', '', '', ''])

        # Percentages
        data.append(['Coverage Percentages', 'TOD Areas', 'Non-TOD Areas', ''])
        data.append(['No Coverage %',
                    f"{tod['no_coverage']/tod['total_roads']*100:.1f}%",
                    f"{non_tod['no_coverage']/non_tod['total_roads']*100:.1f}%", ''])
        data.append(['One-Side %',
                    f"{tod['one_side']/tod['total_roads']*100:.1f}%",
                    f"{non_tod['one_side']/non_tod['total_roads']*100:.1f}%", ''])
        data.append(['Both-Sides %',
                    f"{tod['both_sides']/tod['total_roads']*100:.1f}%",
                    f"{non_tod['both_sides']/non_tod['total_roads']*100:.1f}%", ''])
        data.append(['', '', '', ''])

        # Key insight
        data.append(['KEY INSIGHT:', '', '', ''])
        data.append([f'TOD areas have {tod["any_coverage_pct"]/non_tod["any_coverage_pct"]:.1f}x better sidewalk coverage', '', '', ''])
        data.append(['This indicates targeted investment near transit stations', '', '', ''])

        df = pd.DataFrame(data)

        output_file = self.output_dir / "2_TOD_COMPARISON.xlsx"
        with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='TOD Comparison', index=False, header=False)

            worksheet = writer.sheets['TOD Comparison']
            worksheet.column_dimensions['A'].width = 45
            worksheet.column_dimensions['B'].width = 20
            worksheet.column_dimensions['C'].width = 20
            worksheet.column_dimensions['D'].width = 20

        print(f"[OK] Created: {output_file}")
        return output_file

    def generate_road_type_analysis(self):
        """
        ONE SHEET: Coverage by road type
        """
        print("\n=== Generating Road Type Analysis Excel (ONE SHEET) ===")

        # Extract road type statistics (it's a list, not a dict)
        road_types = self.county_stats.get('by_road_type', [])

        data = []
        data.append(['SIDEWALK COVERAGE BY ROAD TYPE', '', '', ''])
        data.append(['', '', '', ''])

        if road_types and isinstance(road_types, list):
            data.append(['Road Type', 'Total Roads', 'Coverage %', 'No Coverage Count'])

            # Sort by coverage percentage
            sorted_types = sorted(road_types,
                                key=lambda x: x.get('coverage_pct', 0),
                                reverse=True)

            for stats in sorted_types:
                data.append([
                    stats.get('road_type', 'Unknown'),
                    stats.get('total_count', 0),
                    f"{stats.get('coverage_pct', 0):.1f}%",
                    stats.get('no_coverage_count', 0)
                ])
        else:
            data.append(['Road type data not available in statistics', '', '', ''])
            data.append(['', '', '', ''])
            data.append(['Summary statistics only:', '', '', ''])
            data.append(['Total Roads', self.county_stats['total_roads'], '', ''])
            data.append(['County-Wide Coverage',
                        f"{self.county_stats['coverage_percentages']['any_coverage_pct']}%", '', ''])

        df = pd.DataFrame(data)

        output_file = self.output_dir / "3_ROAD_TYPE_ANALYSIS.xlsx"
        with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='Road Type Analysis', index=False, header=False)

            worksheet = writer.sheets['Road Type Analysis']
            worksheet.column_dimensions['A'].width = 35
            worksheet.column_dimensions['B'].width = 20
            worksheet.column_dimensions['C'].width = 20
            worksheet.column_dimensions['D'].width = 25

        print(f"[OK] Created: {output_file}")
        return output_file

    def generate_area_analysis(self):
        """
        ONE SHEET: Coverage by area (acres)
        """
        print("\n=== Generating Area Analysis Excel (ONE SHEET) ===")

        data = []
        data.append(['SIDEWALK COVERAGE BY AREA (ACRES)', '', '', ''])
        data.append(['', '', '', ''])

        # County-wide area statistics
        total_road_area = self.county_stats.get('total_road_area_acres', 0)

        data.append(['County-Wide Road Network', '', '', ''])
        data.append(['Total Road Area', f'{total_road_area:,.1f} acres', '', ''])
        data.append(['Total Roads', self.county_stats['total_roads'], '', ''])
        data.append(['', '', '', ''])

        # Coverage by type
        data.append(['Coverage Type', 'Road Count', 'Percentage', ''])
        data.append(['No Coverage',
                    self.county_stats['coverage_counts']['no_coverage'],
                    f"{self.county_stats['coverage_counts']['no_coverage']/self.county_stats['total_roads']*100:.1f}%", ''])
        data.append(['One-Side Coverage',
                    self.county_stats['coverage_counts']['one_side'],
                    f"{self.county_stats['coverage_counts']['one_side']/self.county_stats['total_roads']*100:.1f}%", ''])
        data.append(['Both-Sides Coverage',
                    self.county_stats['coverage_counts']['both_sides'],
                    f"{self.county_stats['coverage_counts']['both_sides']/self.county_stats['total_roads']*100:.1f}%", ''])
        data.append(['', '', '', ''])

        # TOD area statistics
        data.append(['Transit-Oriented Development (TOD) Areas', '', '', ''])
        data.append(['TOD Coverage',
                    f"{self.tod_stats['tod_statistics']['any_coverage_pct']:.1f}%", '', ''])
        data.append(['TOD Roads', self.tod_stats['tod_statistics']['total_roads'], '', ''])
        data.append(['', '', '', ''])

        data.append(['Non-TOD Areas', '', '', ''])
        data.append(['Non-TOD Coverage',
                    f"{self.tod_stats['non_tod_statistics']['any_coverage_pct']:.1f}%", '', ''])
        data.append(['Non-TOD Roads', self.tod_stats['non_tod_statistics']['total_roads'], '', ''])

        df = pd.DataFrame(data)

        output_file = self.output_dir / "4_AREA_ANALYSIS.xlsx"
        with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='Area Analysis', index=False, header=False)

            worksheet = writer.sheets['Area Analysis']
            worksheet.column_dimensions['A'].width = 40
            worksheet.column_dimensions['B'].width = 25
            worksheet.column_dimensions['C'].width = 20
            worksheet.column_dimensions['D'].width = 20

        print(f"[OK] Created: {output_file}")
        return output_file

    def generate_all_reports(self):
        """Generate all four Druck-compliant Excel reports (ONE SHEET each)"""
        print("\n" + "="*80)
        print("GENERATING DRUCK-COMPLIANT EXCEL REPORTS")
        print("CRITICAL: Each file has ONE SHEET ONLY")
        print("="*80)

        files = []
        files.append(self.generate_executive_summary())
        files.append(self.generate_tod_comparison())
        files.append(self.generate_road_type_analysis())
        files.append(self.generate_area_analysis())

        print("\n" + "="*80)
        print("[OK] EXCEL REPORT GENERATION COMPLETE")
        print("="*80)
        print(f"\nGenerated {len(files)} Excel files (ONE SHEET each):")
        for f in files:
            print(f"  • {f.name}")
        print(f"\nLocation: {self.output_dir}")
        print("\nDruck compliance: [OK] ONE SHEET per file")
        print("Answer placement: [OK] Front and center in Executive Summary")

        return files


if __name__ == "__main__":
    base_dir = Path(__file__).parent.parent.parent
    generator = ExcelReportGenerator(base_dir)
    generator.generate_all_reports()
