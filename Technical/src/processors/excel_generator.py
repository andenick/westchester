"""
Druck-Compliant Excel File Generator

Generates Excel files following Druck standards:
- One sheet per file (no multi-tab workbooks)
- Machine-readable columns (no merged cells, proper headers)
- Professional black & white formatting
- Saved to Output/Data/Results/
"""

import pandas as pd
from pathlib import Path
from typing import Dict, List, Optional
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill


class DruckExcelGenerator:
    """Generate Druck-compliant Excel files"""
    
    def __init__(self, output_dir: Optional[str] = None):
        """
        Initialize Excel generator
        
        Args:
            output_dir: Directory to save Excel files (default: ../../Output/Data/Results/)
        """
        if output_dir is None:
            # Navigate from src/processors/ to Output/Data/Results/
            base_path = Path(__file__).parent.parent.parent.parent / "Output" / "Data" / "Results"
        else:
            base_path = Path(output_dir)
        
        self.output_dir = base_path
        self.output_dir.mkdir(parents=True, exist_ok=True)
    
    def create_excel_file(
        self,
        data: pd.DataFrame,
        filename: str,
        sheet_name: str = "Data",
        add_timestamp: bool = True
    ) -> Path:
        """
        Create a Druck-compliant Excel file
        
        Args:
            data: DataFrame to save
            filename: Base filename (will add .xlsx)
            sheet_name: Name for the single sheet
            add_timestamp: Whether to add [YYYY.MM.DD] prefix to filename
            
        Returns:
            Path to created file
        """
        # Add timestamp prefix if requested (Druck standard)
        if add_timestamp:
            timestamp = datetime.now().strftime("[%Y.%m.%d]")
            filename = f"{timestamp} {filename}"
        
        # Ensure .xlsx extension
        if not filename.endswith('.xlsx'):
            filename += '.xlsx'
        
        filepath = self.output_dir / filename
        
        # Write to Excel with openpyxl engine for formatting
        with pd.ExcelWriter(filepath, engine='openpyxl') as writer:
            data.to_excel(writer, sheet_name=sheet_name, index=False)
            
            # Get the worksheet
            workbook = writer.book
            worksheet = workbook[sheet_name]
            
            # Apply Druck formatting
            self._apply_formatting(worksheet, data)
        
        print(f"[SUCCESS] Created Excel file: {filepath}")
        return filepath
    
    def _apply_formatting(self, worksheet, data: pd.DataFrame):
        """
        Apply Druck-compliant formatting to worksheet
        
        Args:
            worksheet: openpyxl worksheet object
            data: Original DataFrame for dimensions
        """
        # Font for headers (bold)
        header_font = Font(name='Calibri', size=11, bold=True, color='000000')
        
        # Font for data (regular)
        data_font = Font(name='Calibri', size=11, color='000000')
        
        # Alignment (left for text, right for numbers)
        left_align = Alignment(horizontal='left', vertical='top')
        right_align = Alignment(horizontal='right', vertical='top')
        
        # Border style (thin black lines)
        thin_border = Border(
            left=Side(style='thin', color='000000'),
            right=Side(style='thin', color='000000'),
            top=Side(style='thin', color='000000'),
            bottom=Side(style='thin', color='000000')
        )
        
        # Header fill (light gray)
        header_fill = PatternFill(start_color='D9D9D9', end_color='D9D9D9', fill_type='solid')
        
        # Format header row (row 1)
        for cell in worksheet[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = left_align
            cell.border = thin_border
        
        # Format data rows
        for row_idx, row in enumerate(worksheet.iter_rows(min_row=2, max_row=len(data) + 1), start=2):
            for col_idx, cell in enumerate(row, start=1):
                cell.font = data_font
                cell.border = thin_border
                
                # Determine alignment based on data type
                col_name = worksheet.cell(row=1, column=col_idx).value
                if col_name in data.columns:
                    dtype = data[col_name].dtype
                    if dtype in ['int64', 'float64', 'int32', 'float32']:
                        cell.alignment = right_align
                    else:
                        cell.alignment = left_align
                else:
                    cell.alignment = left_align
        
        # Auto-adjust column widths
        for column in worksheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            
            for cell in column:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            
            # Set width (with min/max constraints)
            adjusted_width = min(max(max_length + 2, 10), 50)
            worksheet.column_dimensions[column_letter].width = adjusted_width
        
        # Freeze top row (header)
        worksheet.freeze_panes = 'A2'
    
    def create_from_dict_list(
        self,
        data: List[Dict],
        filename: str,
        sheet_name: str = "Data"
    ) -> Path:
        """
        Create Excel file from list of dictionaries
        
        Args:
            data: List of dictionaries (will be converted to DataFrame)
            filename: Base filename
            sheet_name: Sheet name
            
        Returns:
            Path to created file
        """
        df = pd.DataFrame(data)
        return self.create_excel_file(df, filename, sheet_name)
    
    def create_from_json(
        self,
        json_path: str,
        filename: str,
        sheet_name: str = "Data"
    ) -> Path:
        """
        Create Excel file from JSON file
        
        Args:
            json_path: Path to JSON file
            filename: Base filename for output
            sheet_name: Sheet name
            
        Returns:
            Path to created file
        """
        df = pd.read_json(json_path)
        return self.create_excel_file(df, filename, sheet_name)
    
    def create_from_csv(
        self,
        csv_path: str,
        filename: str,
        sheet_name: str = "Data"
    ) -> Path:
        """
        Create Excel file from CSV file
        
        Args:
            csv_path: Path to CSV file
            filename: Base filename for output
            sheet_name: Sheet name
            
        Returns:
            Path to created file
        """
        df = pd.read_csv(csv_path)
        return self.create_excel_file(df, filename, sheet_name)


def main():
    """Demo/test of Excel generator"""
    print("="*60)
    print("Druck-Compliant Excel Generator - Test")
    print("="*60)
    print()
    
    # Create sample data
    sample_data = pd.DataFrame({
        'Municipality': ['Yonkers', 'New Rochelle', 'Mount Vernon', 'White Plains'],
        'Population': [211569, 79446, 73893, 59559],
        'Median Income': [68976, 88577, 58482, 93171],
        'Area (sq mi)': [20.3, 13.2, 4.4, 10.0]
    })
    
    generator = DruckExcelGenerator()
    filepath = generator.create_excel_file(
        sample_data,
        "westchester_municipalities_sample",
        sheet_name="Municipalities"
    )
    
    print()
    print("[SUCCESS] Sample Excel file created successfully")
    print(f"  Location: {filepath}")
    print()
    print("File follows Druck standards:")
    print("  - One sheet per file [YES]")
    print("  - Machine-readable columns [YES]")
    print("  - Professional formatting [YES]")
    print("  - Timestamped filename [YES]")
    print("="*60)


if __name__ == "__main__":
    main()

